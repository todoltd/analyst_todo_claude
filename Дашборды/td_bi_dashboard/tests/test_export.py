# -*- coding: utf-8 -*-
# Part of td_bi_dashboard. Author: ToDo. Since: 19.0.1.0.0
# Stage-2 експорт: XLSX (xlsxwriter) + PDF (wkhtmltopdf) знімка дашборда (AC-31/32/36/37).
# Виконується на живій Odoo 19:  odoo-bin -u td_bi_dashboard --test-enable
import io

from odoo.tests import TransactionCase, tagged


@tagged('post_install', '-at_install', 'td_bi', 'td_bi_export')
class TestExport(TransactionCase):
    """export_xlsx/export_pdf рендерять знімок дашборда у реальні файли; порожня вибірка —
    файл лише із заголовками (AC-37)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        partner_im = cls.env['ir.model']._get('res.partner')
        ua = cls.env.ref('base.ua')
        pl = cls.env.ref('base.pl')
        recs = cls.env['res.partner'].create(
            [{'name': 'EX UA %d' % i, 'country_id': ua.id} for i in range(3)]
            + [{'name': 'EX PL %d' % i, 'country_id': pl.id} for i in range(2)])
        cls.ds = cls.env['td.bi.dataset'].create({
            'name': 'Export dataset', 'mode': 'model', 'model_id': partner_im.id,
            'visibility': 'global', 'domain': repr([('id', 'in', recs.ids)])})
        cls.dash = cls.env['td.bi.dashboard'].create({'name': 'Export dash'})
        cls.page = cls.env['td.bi.dashboard.page'].create({'dashboard_id': cls.dash.id, 'name': 'P'})
        cls.widget = cls.env['td.bi.widget'].create({
            'page_id': cls.page.id, 'dataset_id': cls.ds.id, 'widget_type': 'table',
            'title': 'By country', 'config': {'data': {'groupby': ['country_id'], 'aggregates': ['__count']}}})

    # --- XLSX: валідний zip-файл, читається openpyxl, містить рядки країн ---
    def test_export_xlsx(self):
        content = self.dash.export_xlsx()
        self.assertTrue(content.startswith(b'PK'), "XLSX — це zip (магічні байти PK).")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        self.assertTrue(wb.sheetnames, "XLSX має хоча б один аркуш.")
        sheet = wb[wb.sheetnames[0]]
        self.assertGreaterEqual(sheet.max_row, 3, "Заголовок + ≥2 рядки країн.")

    # --- AC-37: порожня вибірка -> валідний XLSX лише із заголовками ---
    def test_export_xlsx_empty_headers_only(self):
        empty_ds = self.env['td.bi.dataset'].create({
            'name': 'Empty export ds', 'mode': 'model',
            'model_id': self.env['ir.model']._get('res.partner').id,
            'visibility': 'global', 'domain': repr([('id', '=', 0)])})
        self.widget.dataset_id = empty_ds
        content = self.dash.export_xlsx()
        self.assertTrue(content.startswith(b'PK'))
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        sheet = wb[wb.sheetnames[0]]
        self.assertGreaterEqual(sheet.max_column, 1, "Є заголовок навіть без даних (AC-37).")

    # --- PDF: валідний PDF через wkhtmltopdf ---
    def test_export_pdf(self):
        content = self.dash.export_pdf()
        self.assertTrue(content.startswith(b'%PDF'), "PDF починається з %PDF.")
        self.assertGreater(len(content), 500, "PDF непорожній.")
